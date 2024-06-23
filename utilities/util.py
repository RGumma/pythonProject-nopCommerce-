import openpyxl

def getrowcount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

def getcolcount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def readdata(file, sheetName,rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writedata(file, data, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value
    workbook.save(file)


